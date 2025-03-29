import openpyxl
from openpyxl.styles import Font
from database.database import get_users_by_poll_id
from database.models import Question, QuestionResponse, PollResponse, User, Poll
from sqlalchemy.orm import Session
from sqlalchemy import func
from typing import List, Optional
import logging
from io import BytesIO


import logging
from io import BytesIO
from typing import Optional
from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy.orm import Session
from sqlalchemy import func

def generate_excel_report(db: Session, poll_id: int) -> BytesIO:
    """
    Generates an Excel report for a given poll.

    Args:
        db: SQLAlchemy Session.
        poll_id: The ID of the poll.

    Returns:
        BytesIO: In-memory Excel file.
    """
    # Create a new workbook
    workbook = openpyxl.Workbook()

    # Remove default sheet
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    # Log the start of the process
    logging.info(f"Generating Excel report for poll ID {poll_id}")

    # Create a new sheet for poll description
    description_sheet = workbook.create_sheet("Poll Description", 0)

    # Get poll data
    poll: Optional[Poll] = db.query(Poll).filter(Poll.id == poll_id).first()
    if not poll:
        logging.warning(f"Poll with id {poll_id} not found")
        return BytesIO()

    # Log poll details
    logging.info(f"Poll found: Title='{poll.title}', Description='{poll.description}'")

    # Write poll data to the sheet
    description_sheet["A1"] = "Название опроса"
    description_sheet["B1"] = poll.title
    description_sheet["A2"] = "Описание"
    description_sheet["B2"] = poll.description
    description_sheet["A3"] = "Количество вопросов"
    questions = db.query(Question).filter(Question.poll_id == poll_id).order_by(Question.order).all()
    description_sheet["B3"] = len(questions)

    row_num = 4
    for question in questions:
        description_sheet[f"A{row_num}"] = f"Вопрос {question.order}"
        description_sheet[f"B{row_num}"] = question.text
        row_num += 1
        description_sheet[f"A{row_num}"] = "Варианты ответов"
        description_sheet[f"B{row_num}"] = ", ".join(question.options)
        row_num += 1
        description_sheet[f"A{row_num}"] = "Правильные ответы"
        description_sheet[f"B{row_num}"] = ", ".join(question.correct_answers)
        row_num += 2

    # Log the number of questions retrieved
    logging.info(f"Retrieved {len(questions)} questions for poll ID {poll_id}")

    # Create a new sheet for poll results
    sheet = workbook.create_sheet("Poll Results")

    # Define headers
    headers = ["Имя и Фамилия", "Username", "Email", "Итоговый балл"]
    for question in questions:
        headers.append(f"Ответ {question.order}")
        headers.append(f"Правильный ответ {question.order}")

    # Write headers to the sheet
    for col_num, column_title in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.font = Font(bold=True)
        cell.value = column_title

    # Log the creation of headers
    logging.info(f"Headers created for {len(questions)} questions")

    # Get users who participated in the poll
    user_ids = get_users_by_poll_id(db, poll_id, True)

    # Log the number of users retrieved
    logging.info(f"Retrieved {len(user_ids)} users who participated in poll ID {poll_id}")

    # Write data for each user
    row_num = 2
    for user_id in user_ids:
        user: Optional[User] = db.query(User).filter(User.telegram_id == user_id).first()
        if not user:
            logging.warning(f"User with telegram_id {user_id} not found")
            continue

        # Log user details
        logging.info(f"Processing user: ID={user_id}, Name='{user.first_name} {user.last_name or ''}', Username='{user.username}'")

        # Get user's poll response
        poll_response: Optional[PollResponse] = db.query(PollResponse).filter(
            PollResponse.poll_id == poll_id,
            PollResponse.user_id == user_id
        ).first()
        if not poll_response:
            logging.warning(f"PollResponse not found for user {user_id} and poll {poll_id}")
            continue

        # Calculate total score
        total_score = db.query(func.sum(QuestionResponse.score)).filter(
            QuestionResponse.poll_response_id == poll_response.id
        ).scalar() or 0.0

        # Log user's total score
        logging.info(f"User {user_id} total score: {total_score}")

        user_data = [f"{user.first_name} {user.last_name if user.last_name else ''}", user.username, user.email, total_score]

        # Get user's question responses
        for question in questions:
            question_response: Optional[QuestionResponse] = db.query(QuestionResponse).filter(
                QuestionResponse.poll_response_id == poll_response.id,
                QuestionResponse.question_id == question.id
            ).first()

            if question_response:
                user_answer = ", ".join(question_response.selected_answers)
                correct_answers = ", ".join(question.correct_answers)
                logging.info(f"User {user_id} answered question {question.order}: '{user_answer}' (Correct: '{correct_answers}')")
            else:
                user_answer = "N/A"
                correct_answers = ", ".join(question.correct_answers)
                logging.warning(f"No answer found for user {user_id} on question {question.order}")

            user_data.append(user_answer)
            user_data.append(correct_answers)

        # Log the final user data being written
        logging.info(f"Writing user data to Excel: {user_data}")

        # Write user data to the sheet
        for col_num, cell_value in enumerate(user_data, 1):
            sheet.cell(row=row_num, column=col_num).value = cell_value

        row_num += 1

    # Log the completion of writing all user data
    logging.info(f"All user data written to Excel for poll ID {poll_id}")

    # Create an in-memory file
    excel_file = BytesIO()
    workbook.save(excel_file)
    excel_file.seek(0)

    # Log the successful generation of the Excel file
    logging.info(f"Excel report successfully generated for poll ID {poll_id}")

    return excel_file